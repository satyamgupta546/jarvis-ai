"""
JH Vidhan Sabha Research — 81 Constituencies, 177 Towns
Marts, Bike Showrooms, Banks, Schools, Pharmacy, Optical
Google Maps Places API → Excel (7 sheets)
"""

import json
import csv
import time
import urllib.request
import urllib.parse
import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path
from collections import defaultdict

API_KEY = "AIzaSyChnRGxZ30DZRu2qc68oN6YyJZy5zXpA2A"
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# 81 Vidhan Sabha Constituencies with towns
CONSTITUENCIES = {
    "Rajmahal": {"district": "Sahibganj", "towns": ["Rajmahal", "Udhwa"]},
    "Borio": {"district": "Sahibganj", "towns": ["Borio", "Taljhari"]},
    "Barhait": {"district": "Sahibganj", "towns": ["Barhait", "Pathna"]},
    "Litipara": {"district": "Pakur", "towns": ["Litipara", "Amrapara"]},
    "Pakur": {"district": "Pakur", "towns": ["Pakur", "Hiranpur"]},
    "Maheshpur": {"district": "Pakur", "towns": ["Maheshpur"]},
    "Sikaripara": {"district": "Dumka", "towns": ["Sikaripara", "Masalia"]},
    "Dumka": {"district": "Dumka", "towns": ["Dumka", "Hansdiha"]},
    "Jama": {"district": "Dumka", "towns": ["Jama", "Kundahit", "Gopikandar"]},
    "Nala": {"district": "Jamtara", "towns": ["Nala", "Narayanpur"]},
    "Jamtara": {"district": "Jamtara", "towns": ["Jamtara", "Fatehpur", "Kundhit"]},
    "Jarmundi": {"district": "Godda", "towns": ["Jarmundi", "Sundarpahari"]},
    "Poreyahat": {"district": "Godda", "towns": ["Poreyahat", "Pathargama"]},
    "Godda": {"district": "Godda", "towns": ["Godda", "Boarijor"]},
    "Mahagama": {"district": "Godda", "towns": ["Mahagama", "Meharma"]},
    "Madhupur": {"district": "Deoghar", "towns": ["Madhupur", "Karon"]},
    "Sarath": {"district": "Deoghar", "towns": ["Sarath", "Mohanpur"]},
    "Deoghar": {"district": "Deoghar", "towns": ["Deoghar", "Jasidih"]},
    "Kodarma": {"district": "Koderma", "towns": ["Koderma", "Jhumri Telaiya", "Domchanch"]},
    "Barkatha": {"district": "Hazaribagh", "towns": ["Barkatha", "Katkamsandi"]},
    "Barhi": {"district": "Hazaribagh", "towns": ["Barhi", "Ichak"]},
    "Mandu": {"district": "Hazaribagh", "towns": ["Mandu", "Daru"]},
    "Hazaribagh": {"district": "Hazaribagh", "towns": ["Hazaribagh", "Keredari"]},
    "Simaria": {"district": "Chatra", "towns": ["Simaria", "Tandwa"]},
    "Chatra": {"district": "Chatra", "towns": ["Chatra", "Hunterganj", "Pratappur"]},
    "Barkagaon": {"district": "Ramgarh", "towns": ["Barkagaon", "Patratu", "Kuju"]},
    "Ramgarh": {"district": "Ramgarh", "towns": ["Ramgarh", "Chitarpur", "Gola"]},
    "Dhanwar": {"district": "Giridih", "towns": ["Dhanwar"]},
    "Bagodar": {"district": "Giridih", "towns": ["Bagodar", "Birni"]},
    "Jamua": {"district": "Giridih", "towns": ["Jamua", "Sariya"]},
    "Gandey": {"district": "Giridih", "towns": ["Gandey", "Bengabad"]},
    "Giridih": {"district": "Giridih", "towns": ["Giridih", "Tisri"]},
    "Dumri": {"district": "Giridih", "towns": ["Dumri", "Pirtand"]},
    "Tundi": {"district": "Giridih", "towns": ["Tundi", "Deori"]},
    "Gomia": {"district": "Bokaro", "towns": ["Gomia", "Bermo", "Phusro"]},
    "Bermo": {"district": "Bokaro", "towns": ["Bermo", "Jaridih", "Bokaro Thermal"]},
    "Bokaro": {"district": "Bokaro", "towns": ["Bokaro Steel City", "Chas", "Sector 4"]},
    "Chandankiyari": {"district": "Dhanbad", "towns": ["Chandankiyari", "Baliapur"]},
    "Sindri": {"district": "Dhanbad", "towns": ["Sindri", "Topchanchi"]},
    "Nirsa": {"district": "Dhanbad", "towns": ["Nirsa", "Chirkunda", "Mugma"]},
    "Dhanbad": {"district": "Dhanbad", "towns": ["Dhanbad", "Bank More", "Hirapur"]},
    "Jharia": {"district": "Dhanbad", "towns": ["Jharia", "Katras", "Sijua", "Kusunda"]},
    "Baghmara": {"district": "East Singhbhum", "towns": ["Baghmara", "Dumaria"]},
    "Baharagora": {"district": "East Singhbhum", "towns": ["Baharagora", "Chakulia"]},
    "Ghatsila": {"district": "East Singhbhum", "towns": ["Ghatsila", "Musabani", "Dhalbhumgarh"]},
    "Potka": {"district": "East Singhbhum", "towns": ["Potka", "Patamda"]},
    "Jugsalai": {"district": "East Singhbhum", "towns": ["Jugsalai", "Mango"]},
    "Jamshedpur East": {"district": "East Singhbhum", "towns": ["Jamshedpur", "Bistupur", "Sakchi"]},
    "Jamshedpur West": {"district": "East Singhbhum", "towns": ["Jamshedpur", "Adityapur", "Gamharia"]},
    "Ichagarh": {"district": "Seraikela-Kharsawan", "towns": ["Ichagarh", "Chandil"]},
    "Seraikella": {"district": "Seraikela-Kharsawan", "towns": ["Seraikela", "Kuchai"]},
    "Kharsawan": {"district": "Seraikela-Kharsawan", "towns": ["Kharsawan", "Rajnagar"]},
    "Chaibasa": {"district": "West Singhbhum", "towns": ["Chaibasa"]},
    "Majhgaon": {"district": "West Singhbhum", "towns": ["Majhgaon", "Noamundi"]},
    "Jaganathpur": {"district": "West Singhbhum", "towns": ["Jagannathpur", "Manoharpur"]},
    "Manoharpur": {"district": "West Singhbhum", "towns": ["Manoharpur", "Sonua"]},
    "Chakradharpur": {"district": "West Singhbhum", "towns": ["Chakradharpur", "Bandgaon"]},
    "Tamar": {"district": "Ranchi", "towns": ["Tamar", "Bundu"]},
    "Silli": {"district": "Ranchi", "towns": ["Silli", "Sonahatu"]},
    "Khijri": {"district": "Ranchi", "towns": ["Khijri", "Murhu"]},
    "Ranchi": {"district": "Ranchi", "towns": ["Ranchi", "Main Road", "Lalpur"]},
    "Hatia": {"district": "Ranchi", "towns": ["Hatia", "Dhurwa", "Argora"]},
    "Kanke": {"district": "Ranchi", "towns": ["Kanke", "Ratu", "Namkum"]},
    "Torpa": {"district": "Khunti", "towns": ["Torpa", "Arki"]},
    "Khunti": {"district": "Khunti", "towns": ["Khunti", "Karra"]},
    "Mandar": {"district": "Lohardaga", "towns": ["Mandar", "Kisko"]},
    "Lohardaga": {"district": "Lohardaga", "towns": ["Lohardaga", "Senha"]},
    "Sisai": {"district": "Gumla", "towns": ["Sisai", "Chainpur"]},
    "Gumla": {"district": "Gumla", "towns": ["Gumla", "Bishunpur"]},
    "Bishunpur": {"district": "Gumla", "towns": ["Bishunpur", "Raidih"]},
    "Simdega": {"district": "Simdega", "towns": ["Simdega", "Thethaitangar"]},
    "Kolebira": {"district": "Simdega", "towns": ["Kolebira", "Bano"]},
    "Manika": {"district": "Latehar", "towns": ["Manika", "Balumath"]},
    "Latehar": {"district": "Latehar", "towns": ["Latehar", "Chandwa"]},
    "Panki": {"district": "Palamu", "towns": ["Panki", "Patan"]},
    "Daltonganj": {"district": "Palamu", "towns": ["Daltonganj", "Medininagar"]},
    "Bishrampur": {"district": "Palamu", "towns": ["Bishrampur"]},
    "Chhatarpur": {"district": "Palamu", "towns": ["Chhatarpur", "Satbarwa"]},
    "Hussainabad": {"district": "Palamu", "towns": ["Hussainabad", "Chainpur"]},
    "Garhwa": {"district": "Garhwa", "towns": ["Garhwa", "Nagar Untari"]},
    "Bhawanathpur": {"district": "Garhwa", "towns": ["Bhawanathpur", "Ranka"]},
}

SEARCHES = {
    "Mart": ["Vishal Mega Mart", "D-Mart", "Reliance Smart", "Big Bazaar", "More Supermarket", "Spencer's", "Supermarket", "Departmental Store"],
    "Bike Showroom": ["Honda showroom", "Hero showroom", "TVS showroom", "Bajaj showroom", "Suzuki showroom", "Royal Enfield showroom", "Yamaha showroom", "KTM showroom"],
    "Bank": ["HDFC Bank", "ICICI Bank", "SBI Bank", "Axis Bank", "PNB Bank", "Bank of Baroda", "Canara Bank", "Union Bank", "Kotak Mahindra Bank"],
    "School": ["DAV School", "DPS School", "Kendriya Vidyalaya"],
    "Pharmacy": ["Apollo Pharmacy", "MedPlus Pharmacy"],
    "Optical": ["Lenskart", "Titan Eye Plus"],
}

BRAND_MAP = {
    "Bike Showroom": ["Honda", "Hero", "TVS", "Bajaj", "Suzuki", "Royal Enfield", "Yamaha", "KTM"],
    "Bank": ["HDFC", "ICICI", "SBI", "Axis", "PNB", "Bank of Baroda", "Canara", "Union", "Kotak"],
    "School": ["DAV", "DPS", "Kendriya Vidyalaya", "KV"],
    "Pharmacy": ["Apollo", "MedPlus"],
    "Optical": ["Lenskart", "Titan Eye"],
}


def search_places(query, town):
    full_query = f"{query} in {town}, Jharkhand, India"
    url = f"https://maps.googleapis.com/maps/api/place/textsearch/json?query={urllib.parse.quote(full_query)}&key={API_KEY}"
    try:
        resp = json.loads(urllib.request.urlopen(urllib.request.Request(url), timeout=15).read())
        results = []
        for p in resp.get("results", []):
            addr = p.get("formatted_address", "")
            area = addr.split(",")[0].strip() if "," in addr else ""
            results.append({
                "name": p.get("name", ""),
                "area": area,
                "address": addr,
                "rating": p.get("rating", ""),
                "total_ratings": p.get("user_ratings_total", ""),
                "maps_link": f"https://www.google.com/maps/place/?q=place_id:{p.get('place_id', '')}",
            })
        return results
    except:
        return []


def detect_brand(name, category):
    brands = BRAND_MAP.get(category, [])
    for b in brands:
        if b.lower() in name.lower():
            return b
    return ""


def main():
    total_towns = sum(len(v["towns"]) for v in CONSTITUENCIES.values())
    print(f"\n{'='*60}")
    print(f"  JH VIDHAN SABHA RESEARCH")
    print(f"  {len(CONSTITUENCIES)} Constituencies | {total_towns} Towns")
    print(f"  6 Categories | Google Maps Places API")
    print(f"{'='*60}\n")

    all_data = {cat: [] for cat in SEARCHES}
    seen = {cat: set() for cat in SEARCHES}

    for const_name, const_data in CONSTITUENCIES.items():
        district = const_data["district"]
        towns = const_data["towns"]
        print(f"\n[{const_name}] ({district}) — {len(towns)} towns")

        for town in towns:
            for category, queries in SEARCHES.items():
                town_count = 0
                for query in queries:
                    results = search_places(query, town)
                    for r in results:
                        key = f"{r['name']}|{town}"
                        if key not in seen[category]:
                            seen[category].add(key)
                            r["constituency"] = const_name
                            r["district"] = district
                            r["town"] = town
                            r["category"] = category
                            r["brand"] = detect_brand(r["name"], category)
                            all_data[category].append(r)
                            town_count += 1
                    time.sleep(0.2)
                if town_count:
                    print(f"  {town}: {town_count} {category}")

    # Save CSVs
    for cat, data in all_data.items():
        if data:
            path = OUTPUT_DIR / f"vs_{cat.lower().replace(' ', '_')}.csv"
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=data[0].keys())
                writer.writeheader()
                writer.writerows(data)
            print(f"\n[CSV] {cat}: {len(data)} rows → {path.name}")

    # Build Excel
    print("\n[Excel] Building...")
    wb = openpyxl.Workbook()
    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    headers = ["Constituency", "District", "Town", "Area", "Name", "Type", "Brand", "Address", "Rating", "Total Ratings", "Google Maps"]

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    sum_headers = ["Constituency", "District", "Towns", "Marts", "Showrooms", "Banks", "Schools", "Pharmacy", "Optical",
                   "Honda", "Hero", "TVS", "Bajaj", "RE", "HDFC", "ICICI", "SBI", "Grand Total"]
    ws.append(sum_headers)
    for cell in ws[1]:
        cell.font = hfont; cell.fill = hfill

    for const_name, const_data in CONSTITUENCIES.items():
        district = const_data["district"]
        towns_str = ", ".join(const_data["towns"])
        counts = {}
        brand_counts = defaultdict(int)
        for cat, data in all_data.items():
            cat_data = [r for r in data if r["constituency"] == const_name]
            counts[cat] = len(cat_data)
            for r in cat_data:
                if r["brand"]:
                    brand_counts[r["brand"]] += 1

        total = sum(counts.values())
        ws.append([const_name, district, towns_str, counts.get("Mart", 0), counts.get("Bike Showroom", 0),
                   counts.get("Bank", 0), counts.get("School", 0), counts.get("Pharmacy", 0), counts.get("Optical", 0),
                   brand_counts.get("Honda", 0), brand_counts.get("Hero", 0), brand_counts.get("TVS", 0),
                   brand_counts.get("Bajaj", 0), brand_counts.get("Royal Enfield", 0),
                   brand_counts.get("HDFC", 0), brand_counts.get("ICICI", 0), brand_counts.get("SBI", 0), total])

    # Data sheets
    for cat, data in all_data.items():
        sheet_name = cat[:31]  # Excel max 31 chars
        ws2 = wb.create_sheet(sheet_name)
        ws2.append(headers)
        for cell in ws2[1]:
            cell.font = hfont; cell.fill = hfill
        for r in sorted(data, key=lambda x: (x["constituency"], x["town"])):
            ws2.append([r["constituency"], r["district"], r["town"], r["area"], r["name"],
                       r["category"], r["brand"], r["address"], r["rating"], r["total_ratings"], r["maps_link"]])

    # Auto width
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    excel_path = "/Users/satyam/Downloads/JH_Vidhan_Sabha_Research.xlsx"
    wb.save(excel_path)

    total_all = sum(len(d) for d in all_data.values())
    print(f"\n{'='*60}")
    print(f"  DONE!")
    print(f"  Constituencies: {len(CONSTITUENCIES)}")
    print(f"  Towns: {total_towns}")
    print(f"  Total places: {total_all}")
    for cat, data in all_data.items():
        print(f"    {cat}: {len(data)}")
    print(f"  Excel: {excel_path}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
